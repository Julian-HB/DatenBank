import os
from openpyxl import load_workbook
import sqlite3
from sqlite3 import Error
import sys
from pathlib import Path

class InsertInTable:

    def __init__(self, table_name, excel_directory, min_row, max_row, min_col, max_col):
        wb = load_workbook(excel_directory)
        ws = wb.worksheets[0]
        self.table = table_name
        self.general_range = ws.iter_rows(min_row, max_row, min_col, max_col)
        self.y_Range = []
        self.x_Range = []
        self.data = []



    def load_data(self):
        for index, (x, y) in enumerate(self.general_range):
            #x.value = x.value.strftime('%Y-%m-%d')
            self.data.append((index, x.value, y.value))

    def database_insert(self):
        # Herstellung der Verbindung zu der Datenbank.
        database_path = os.path.join(r'C:\Users\julia\Documents\tutorials\Praedico\Prognose\database_files\database_db_files', 'database.db')

        connection = sqlite3.connect(database_path)
        # Cusorobject wird erstellt. Der Cursor dient zur Tabellenmanipulation.
        cursor = connection.cursor()
        # Daten werden in Tabelle Geldmenge eingefügt.
        cursor.executemany(f'INSERT INTO {self.table} VALUES (?, ?, ?)', self.data)
        # Änderungen werden dauerhaft gespeichert.
        connection.commit()
        connection.close()


# Example usage:
current_dir = os.getcwd()
destination_dir = r'C:\Users\julia\Documents\tutorials\Praedico\Prognose\database_files\database_raw_excel_files\Geldmenge_M1.xlsx'

inserter = InsertInTable("Geldmenge_M1", destination_dir, 2, 651, 1, 2)
inserter.load_data()
inserter.database_insert()